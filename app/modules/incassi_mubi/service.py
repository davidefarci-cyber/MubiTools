"""Orchestratore del modulo Incassi Mubi.

Espone `elabora_incassi` — l'unica funzione chiamata dal router — che
coordina le 6 fasi definite in `processor.py`, valida la presenza delle
colonne richieste e salva i file di output (conferimento aggiornato,
report anomalie, nuove righe).

La firma di `elabora_incassi` è parte del contratto pubblico del modulo:
non modificarla senza aggiornare `router.py:128-135`.
"""

import logging
from collections.abc import Callable
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from app.modules.incassi_mubi.excel_reader import (
    COL_IMPORTO_APERTO_VARIANTS,
    COL_MODALITA_PAGAMENTO_VARIANTS,
    COL_NR_BOLLETTA_VARIANTS,
    COL_DATA_PAGAMENTO_VARIANTS,
    _read_excel_smart,
)
from app.shared.excel_mapper import find_column
from app.modules.incassi_mubi.processor import (
    fase1_parse_incassi,
    fase2_join_importo_aperto,
    fase3_piani_rientro,
    fase4_popola_conferimento,
    fase5_calcolo_incassato,
    fase6_ordinamento_controllo,
)
from app.modules.incassi_mubi.validator import _normalize_amount, _validate_all_columns

logger = logging.getLogger(__name__)

RED_FILL = PatternFill(start_color="FFE74C3C", end_color="FFE74C3C", fill_type="solid")


# ─── Salvataggio output ──────────────────────────────────────────

def salva_conferimento(
    df_conferimento: pd.DataFrame,
    anomalie: list[dict],
    output_path: Path,
    template_path: Path | None = None,
) -> None:
    """Salva il DataFrame conferimento aggiornato in un file Excel.

    Evidenzia in rosso le righe con anomalie.
    """
    logger.info("Salvataggio conferimento aggiornato: %s", output_path)

    col_boll = find_column(df_conferimento, COL_NR_BOLLETTA_VARIANTS)
    anomaly_bollette = {a["numero_bolletta"] for a in anomalie}

    # Salva con openpyxl per poter applicare stili
    df_conferimento.to_excel(output_path, index=False, engine="openpyxl")

    wb = load_workbook(output_path)
    ws = wb.active

    if col_boll:
        boll_col_idx = list(df_conferimento.columns).index(col_boll) + 1
        for row_idx in range(2, ws.max_row + 1):
            cell_val = str(ws.cell(row=row_idx, column=boll_col_idx).value or "").strip()
            if cell_val in anomaly_bollette:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = RED_FILL

    wb.save(output_path)
    wb.close()
    logger.info("  File salvato: %d righe", len(df_conferimento))


def salva_report_anomalie(anomalie: list[dict], output_path: Path) -> None:
    """Salva il report anomalie in un file Excel."""
    if not anomalie:
        return
    df = pd.DataFrame(anomalie)
    df.to_excel(output_path, index=False, engine="openpyxl")
    logger.info("Report anomalie salvato: %d righe -> %s", len(df), output_path)


def salva_nuove_righe(df_nuove: pd.DataFrame, output_path: Path) -> None:
    """Salva le nuove righe da aggiungere al conferimento."""
    if df_nuove.empty:
        return
    df_nuove.to_excel(output_path, index=False, engine="openpyxl")
    logger.info("Nuove righe salvate: %d -> %s", len(df_nuove), output_path)


# ─── Orchestratore ────────────────────────────────────────────────

def elabora_incassi(
    file_incassi: Path,
    file_massivo: Path,
    file_conferimento: Path,
    file_piani: Path | None,
    output_dir: Path,
    progress_callback: Callable[[int, str], None] | None = None,
) -> dict:
    """Esegue le 6 fasi di elaborazione e restituisce i risultati.

    Args:
        file_incassi: Path al file .txt esportato da Mubi
        file_massivo: Path al file .xlsx estrazione massiva
        file_conferimento: Path al file .xlsx conferimento
        file_piani: Path al file .xlsx piani di rientro (opzionale)
        output_dir: Directory dove salvare i file output
        progress_callback: Funzione(phase, message) per aggiornare lo stato

    Returns:
        Dizionario con risultati e percorsi file output.
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    def notify(phase: int, msg: str) -> None:
        logger.info("  [Fase %d] %s", phase, msg)
        if progress_callback:
            progress_callback(phase, msg)

    debug_infos: list[dict] = []

    # FASE 1
    notify(1, "Parsing file incassi...")
    df_incassi = fase1_parse_incassi(file_incassi)
    debug_incassi = {
        "file": "incassi",
        "sheets_available": ["—"],
        "sheet_used": "—",
        "columns": list(df_incassi.columns),
        "columns_matched": {},
        "columns_missing": {},
    }
    # Verifica colonne attese nel file incassi
    for label, variants in [
        ("numerofattura", COL_NR_BOLLETTA_VARIANTS),
        ("importoaperto", COL_IMPORTO_APERTO_VARIANTS),
        ("datapagamento", COL_DATA_PAGAMENTO_VARIANTS),
        ("metodopagamento", COL_MODALITA_PAGAMENTO_VARIANTS),
    ]:
        col = find_column(df_incassi, variants)
        if col:
            debug_incassi["columns_matched"][label] = col
        else:
            debug_incassi["columns_missing"][label] = variants
    debug_infos.append(debug_incassi)
    notify(1, f"Completato: {len(df_incassi)} righe")

    # FASE 2
    notify(2, "Join massivo con incassi per ImportoAperto...")
    df_massivo, df_nuove_righe, df_incassi, debug_massivo = fase2_join_importo_aperto(df_incassi, file_massivo)
    debug_infos.append(debug_massivo)
    notify(2, f"Completato: {len(df_nuove_righe)} nuove righe (ImportoAperto > 20)")

    # Carica conferimento con smart reader
    df_conferimento, debug_conf = _read_excel_smart(
        file_conferimento,
        required_variants=[
            COL_NR_BOLLETTA_VARIANTS,
            COL_IMPORTO_APERTO_VARIANTS,
            ["incassato", "importo incassato"],
        ],
        label="conferimento",
    )
    debug_infos.append(debug_conf)

    # Validazione colonne — errore anticipato con dettagli
    validation_errors = _validate_all_columns(debug_infos)
    if validation_errors:
        error_lines = []
        for err in validation_errors:
            error_lines.append(
                f"[{err['file']}] Colonna '{err['colonna_attesa']}' non trovata. "
                f"Foglio usato: '{err['sheet']}' (fogli disponibili: {err['sheets_available']}). "
                f"Colonne nel foglio: {err['colonne_trovate'][:10]}..."
            )
        error_msg = "Validazione colonne fallita:\n" + "\n".join(error_lines)
        logger.error(error_msg)
        raise ValueError(error_msg)

    # FASE 3
    notify(3, "Verifica piani di rientro...")
    df_conferimento, piani_count, debug_piani = fase3_piani_rientro(
        df_conferimento, file_piani
    )
    if debug_piani:
        debug_infos.append(debug_piani)
    notify(3, f"Completato: {piani_count} piani trovati")

    # FASE 4
    notify(4, "Popola colonne Conferimento...")
    df_conferimento = fase4_popola_conferimento(df_conferimento, df_incassi)
    notify(4, "Completato")

    # FASE 5
    notify(5, "Calcolo Incassato...")
    df_conferimento = fase5_calcolo_incassato(df_conferimento)
    notify(5, "Completato")

    # FASE 6
    notify(6, "Ordinamento e controllo...")
    df_conferimento, anomalie = fase6_ordinamento_controllo(df_conferimento)
    notify(6, f"Completato: {len(anomalie)} anomalie")

    # Salva output
    output_conferimento = output_dir / "conferimento_aggiornato.xlsx"
    salva_conferimento(df_conferimento, anomalie, output_conferimento, file_conferimento)

    output_anomalie = output_dir / "report_anomalie.xlsx"
    salva_report_anomalie(anomalie, output_anomalie)

    output_nuove = output_dir / "nuove_righe_conferimento.xlsx"
    salva_nuove_righe(df_nuove_righe, output_nuove)

    # Calcola statistiche
    col_z = find_column(df_conferimento, ["incassato", "importo incassato"])

    fatture_incassate = 0
    if col_z:
        for _, row in df_conferimento.iterrows():
            val = _normalize_amount(row.get(col_z, 0))
            if val > 0:
                fatture_incassate += 1

    results = {
        "total_fatture": len(df_conferimento),
        "fatture_incassate": fatture_incassate,
        "anomalie": len(anomalie),
        "piani_rientro": piani_count,
        "nuove_righe": len(df_nuove_righe),
        "anomalie_detail": anomalie[:100],  # Max 100 per la UI
        "debug_info": debug_infos,
        "files": {
            "conferimento": str(output_conferimento),
            "anomalie": str(output_anomalie) if anomalie else None,
            "nuove_righe": str(output_nuove) if not df_nuove_righe.empty else None,
        },
    }

    logger.info("Elaborazione completata: %s", {
        k: v for k, v in results.items() if k not in ("anomalie_detail", "debug_info", "files")
    })
    return results
