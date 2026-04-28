"""Logica di business per il modulo Connessione.

Sottofunzionalita': Crea Riga FILE A
Legge un file Excel (FILE B), mappa le colonne verso il formato FILE A,
applica trasformazioni sui valori e scrive il risultato come nuovo foglio.
"""

import csv
import io
import logging
import uuid
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from app.shared.excel_mapper import find_column

logger = logging.getLogger(__name__)

# ===== Mappatura colonne FILE B -> FILE A =====
COLUMN_MAP_B_TO_A = {
    "ATTIVITA'": "ATTIVITA'",
    "DATA RICEZIONE": "OGGI",
    "DATA EVASIONE": "",
    "DATA INVIO 150": "",
    "NOTE": "",
    "COD SII": "",
    "VENDITORE": "",
    "RAGSOC": "RAGSOC",
    "CF": "CF",
    "P. IVA": "PIVA",
    "NR_TELEFONO": "NR_TELEFONO",
    "DUG FORNITURA": "DUG FORNITURA",
    "TOPONIMO": "INDIRIZZO FORNITURA",
    "CIVICO": "CIVICO FORNITURA",
    "CAP": "CAP FORNITURA",
    "COMUNE": "LOCALITA FORNITURA",
    "PROV": "PROVINCIA FORNITURA",
    "PDR": "PDR",
    "MATRICOLA": "MATRICOLA",
    "CODICE DL": "",
    "REMI": "REMI",
    "DL": "DISTRIBUTORE",
    "POTENZA": "Potenzialità massima richiesta (in kw)",
    "USO": "Tipo uso",
    "CATEGORIA D'USO": "categoria uso",
    "GG ": "gg utilizzo- classe di prelievo",
    "VOLUME ANNUO": "CONSUMO ANNUO TOTALE STIMATO",
    "DATA APP": "",
    "ORARIO APP": "",
    "COD. APP": "",
    "PREVENTIVO": "",
    "IMPONIBILE": "",
    "ACCETTAZIONE PREV": "",
    "STATO": "RICEZIONE ATTIVITA'",
    "NOTE 2": "",
}

# ===== Mappe valori =====
ATTIVITA_MAP = {
    "A01": "A01 - ATTIVAZIONE SEMPLICE",
    "A40": "A40 - ATTIVAZIONE CON ACCERTAMENTO",
    "PM1": "PM1 - PREVENTIVO MODIFICA IMPIANTO",
    "PN1": "PN1 - PREVENTIVO NUOVO IMPIANTO",
    "E01": "E01 - ESECUZIONE LAVORI",
    "D01": "D01 - DISATTIVAZIONE",
    "PR1": "PR1 - RIMOZIONE",
    "V01": "V01 - VERIFICA GDM",
    "V02": "V02 - VERIFICA PRESSIONE",
    "SM1": "SM1 - SOSPENSIONE MOROSITA'",
    "R01": "R01 - RIATTIVAZIONE MOROSITA'",
    "A02": "A02 - RIATTIVAZIONE P. INTERVENTO",
    "R40": "R40 - RIATTIVAZIONE CON ACCERTAMENTO",
    "V": "V - VOLTURA",
    "MU": "MU - MODIFICA USO",
    "SM2": "SM2 - TAGLIO COLONNA",
    "M02": "M02 - RICHIESTA DATI",
    "M01": "M01 - RETTIFICHE",
    "SPR": "SPR - CAMBIO CONTATORE",
}

CATEGORIA_USO_MAP = {
    "C1": "C1 - RISCALDAMENTO",
    "C2": "C2 - USO COTTURA CIBI E/O PRODUZIONE DI ACQUA CALDA SANITARIA",
    "C3": "C3 - RISCALDAMENTO/USO COTTURA CIBI E/O PRODUZIONE DI ACQUA CALDA SANITARIA",
    "C5": "C5 - USO CONDIZIONAMENTO E RISCALDAMENTO",
    "T1": "T1-1 - USO TECNOLOGICO (ARTIGIANALE-INDUSTRIALE) - 7 GIORNI",
    "T2": "T2-1 - USO TECNOLOGICO/RISCALDAMENTO - 7 GIORNI",
}

GG_MAP = {
    "A": "1 (= 7 gg)",
    "B": "2 (= 6 gg)",
    "C": "3 (= 5 gg)",
}

TIPO_USO_MAP = {
    "domestico": "DOMESTICO",
    "condominio domestico": "CONDOMINIO DOMESTICO",
    "usi diversi": "ALTRI USI",
    "pubblico": "USO PUBBLICO",
}

# Colonne indirizzo: col_a -> nome da cercare in FILE B
_INDIRIZZO_MAP = {
    "TOPONIMO": "INDIRIZZO FORNITURA",
    "CIVICO": "CIVICO FORNITURA",
    "CAP": "CAP FORNITURA",
    "COMUNE": "LOCALITA FORNITURA",
    "PROV": "PROVINCIA FORNITURA",
}


def _clean_val(val: object) -> str:
    """Pulisce un valore: rimuove .0 dai numeri e restituisce stringa vuota per valori nulli."""
    val = str(val).strip()
    if val in ("", "nan", "NaN", "None"):
        return ""
    if val.endswith(".0"):
        try:
            int(val[:-2])
            val = val[:-2]
        except ValueError:
            pass
    return val



def _build_row(df_b: pd.DataFrame, row_idx: int, warnings: set[str]) -> dict:
    """Costruisce una riga FILE A dalla riga row_idx del FILE B."""
    new_row: dict[str, object] = {}

    for col_a, source in COLUMN_MAP_B_TO_A.items():
        if source == "OGGI":
            new_row[col_a] = datetime.now().strftime("%d/%m/%Y")
        elif source == "RICEZIONE ATTIVITA'":
            new_row[col_a] = "RICEZIONE ATTIVITA'"
        elif source == "":
            new_row[col_a] = ""
        else:
            # Indirizzi
            if col_a in _INDIRIZZO_MAP:
                found_col = find_column(df_b, [_INDIRIZZO_MAP[col_a]], mode="substring")
                if found_col:
                    val = _clean_val(df_b[found_col].iloc[row_idx])
                    val = val.upper() if val else ""
                else:
                    val = ""
                    warnings.add(f"Colonna '{_INDIRIZZO_MAP[col_a]}' non trovata per '{col_a}'")
                new_row[col_a] = val

            # PDR
            elif col_a == "PDR":
                if source in df_b.columns:
                    val = _clean_val(df_b[source].iloc[row_idx])
                else:
                    val = ""
                    warnings.add(f"Colonna '{source}' non trovata")
                new_row["PDR"] = val.upper() if val else ""
                new_row["LUNG."] = len(val) if val else ""

            # P. IVA e CF
            elif col_a in ("P. IVA", "CF"):
                if source in df_b.columns:
                    val = _clean_val(df_b[source].iloc[row_idx])
                else:
                    val = ""
                    warnings.add(f"Colonna '{source}' non trovata per '{col_a}'")
                new_row[col_a] = val.upper() if val else ""

            # Altri campi
            else:
                found_col = source if source in df_b.columns else find_column(df_b, [source], mode="substring")
                if found_col:
                    val = _clean_val(df_b[found_col].iloc[row_idx])
                    if val:
                        val = val.strip().upper()
                        if col_a == "ATTIVITA'" and val in ATTIVITA_MAP:
                            val = ATTIVITA_MAP[val].upper()
                        elif col_a == "CATEGORIA D'USO" and val in CATEGORIA_USO_MAP:
                            val = CATEGORIA_USO_MAP[val].upper()
                        elif col_a.strip() == "GG" and val in GG_MAP:
                            val = GG_MAP[val]
                        elif col_a == "USO" and val in TIPO_USO_MAP:
                            val = TIPO_USO_MAP[val].upper()
                    new_row[col_a] = val
                else:
                    new_row[col_a] = ""
                    warnings.add(f"Colonna sorgente '{source}' non trovata per '{col_a}'")

    return new_row


def _get_columns_order() -> list[str]:
    """Restituisce l'ordine delle colonne con LUNG. dopo PDR."""
    columns: list[str] = []
    for k in COLUMN_MAP_B_TO_A:
        columns.append(k)
        if k == "PDR":
            columns.append("LUNG.")
    return columns


def genera_righe_connessione(file_path: Path) -> dict:
    """Legge il file Excel e genera le righe per connessione come dati JSON.

    Args:
        file_path: Path al file Excel sorgente.

    Returns:
        dict con rows_created, columns, rows, warnings.
    """
    logger.info("Genera righe connessione: lettura %s", file_path.name)

    df_b = pd.read_excel(
        file_path,
        dtype={"PDR": str, "PIVA": str, "CF": str},
        keep_default_na=False,
        na_values=[],
    )
    df_b.columns = df_b.columns.str.strip()

    # Prendi solo le righe fino alla prima riga vuota
    valid_rows: list[int] = []
    for idx in range(len(df_b)):
        row = df_b.iloc[idx]
        has_data = any(
            str(v).strip() not in ("", "nan", "NaN", "None")
            for v in row
            if pd.notna(v)
        )
        if not has_data:
            break
        valid_rows.append(idx)

    if not valid_rows:
        raise ValueError("Il file non contiene righe di dati")

    df_b = df_b.iloc[valid_rows].reset_index(drop=True)

    warnings: set[str] = set()
    rows: list[dict] = []
    for i in range(len(df_b)):
        rows.append(_build_row(df_b, i, warnings))

    columns = _get_columns_order()
    # Converti ogni riga-dict in lista ordinata per colonna
    rows_data: list[list[str]] = []
    for row in rows:
        rows_data.append([str(row.get(c, "")) for c in columns])

    logger.info("Genera righe connessione: %d righe create", len(rows_data))

    return {
        "rows_created": len(rows_data),
        "columns": columns,
        "rows": rows_data,
        "warnings": sorted(warnings),
    }


# --- Archiviata: scrittura su foglio Excel (per uso futuro) ---

def crea_riga_file_a(file_path: Path, sheet_name: str = "Riga FILE A") -> dict:
    """Genera righe e le scrive come nuovo foglio Excel nel file sorgente.

    Mantenuta per eventuale riutilizzo futuro in altre sezioni.
    """
    logger.info("Crea Riga FILE A: lettura %s", file_path.name)

    df_b = pd.read_excel(
        file_path,
        dtype={"PDR": str, "PIVA": str, "CF": str},
        keep_default_na=False,
        na_values=[],
    )
    df_b.columns = df_b.columns.str.strip()

    valid_rows: list[int] = []
    for idx in range(len(df_b)):
        row = df_b.iloc[idx]
        has_data = any(
            str(v).strip() not in ("", "nan", "NaN", "None")
            for v in row
            if pd.notna(v)
        )
        if not has_data:
            break
        valid_rows.append(idx)

    if not valid_rows:
        raise ValueError("Il file non contiene righe di dati")

    df_b = df_b.iloc[valid_rows].reset_index(drop=True)

    warnings: set[str] = set()
    rows: list[dict] = []
    for i in range(len(df_b)):
        rows.append(_build_row(df_b, i, warnings))

    df_new = pd.DataFrame(rows)

    book = load_workbook(file_path)
    if sheet_name in book.sheetnames:
        std = book[sheet_name]
        book.remove(std)
    book.save(file_path)
    book.close()

    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:
        df_new.to_excel(writer, sheet_name=sheet_name, index=False)

    logger.info("Crea Riga FILE A: %d righe create in '%s'", len(rows), sheet_name)

    return {
        "rows_created": len(rows),
        "warnings": sorted(warnings),
        "output_path": str(file_path),
    }


def estrai_pod_xml(file_path: Path, pod_list: list[str], upload_dir: Path) -> dict:
    """Legge un file XML in streaming ed estrae i blocchi <DatiPod> matching i POD richiesti.

    Produce un file ZIP contenente:
    - output_{pod}.xml per ogni POD trovato
    - log.txt con lo stato di ogni POD (OK / NON TROVATO)

    Args:
        file_path: Path al file XML sorgente.
        pod_list: Lista di codici POD da cercare.
        upload_dir: Directory dove salvare il file ZIP di output.

    Returns:
        dict con keys: zip_path, found, not_found, total_requested.
    """
    logger.info("Estrai POD XML: lettura %s, %d POD richiesti", file_path.name, len(pod_list))

    pods = set(pod_list)
    found: dict[str, bytes] = {}

    context = ET.iterparse(file_path, events=("start", "end"))
    context = iter(context)
    event, root_elem = next(context)
    root_tag = root_elem.tag

    for event, elem in context:
        if event == "end" and elem.tag == "DatiPod":
            pod_node = elem.find("Pod")
            if pod_node is not None and pod_node.text in pods:
                pod_value = pod_node.text
                new_root = ET.Element(root_tag)
                new_root.append(elem)
                tree = ET.ElementTree(new_root)
                buf = io.BytesIO()
                tree.write(buf, encoding="utf-8", xml_declaration=True)
                found[pod_value] = buf.getvalue()
            root_elem.clear()

    not_found = sorted(pods - set(found.keys()))

    # Costruisce log
    log_lines = [f"{pod} -> OK" for pod in sorted(found.keys())]
    log_lines += [f"{pod} -> NON TROVATO" for pod in not_found]

    # Crea ZIP
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for pod, xml_bytes in found.items():
            zf.writestr(f"output_{pod}.xml", xml_bytes)
        zf.writestr("log.txt", "\n".join(log_lines))
    zip_buf.seek(0)

    job_id = str(uuid.uuid4())
    zip_path = upload_dir / f"pod_output_{job_id}.zip"
    zip_path.write_bytes(zip_buf.read())

    logger.info(
        "Estrai POD XML: %d trovati, %d non trovati -> %s",
        len(found), len(not_found), zip_path.name,
    )

    return {
        "zip_path": str(zip_path),
        "found": sorted(found.keys()),
        "not_found": not_found,
        "total_requested": len(pods),
    }


# ===========================================================================
# S01 Massivo — generazione CSV + XLSX nel formato S01 da Excel pratiche
# ===========================================================================

S01_HEADER_FISSI: tuple[str, str, str, str, str] = (
    "S01",
    "0050",
    "03443420231",
    "DP2032",
    "05779711000",
)

S01_USO_MAP: dict[str, tuple[str, str]] = {
    "ALTRI USI": ("03", "930"),
    "DOMESTICO NON RESIDENTE": ("02", "001"),
    "DOMESTICO RESIDENTE": ("01", "001"),
}

S01_HEADER: list[str] = [
    "COD_SERVIZIO", "COD_FLUSSO", "PIVA_UTENTE", "COD_CONTR_DISP", "PIVA_DISTR",
    "COD_PRAT_UTENTE", "COD_POD", "COGNOME", "NOME", "RAG_SOC", "CF", "PIVA", "TEL",
    "PRESENZA_CLIENTE_NO_TELEGESTITO",
    "TOPONIMO_1", "VIA_1", "CIV_1", "CAP_1", "ISTAT_1", "LOCALITA_1", "PROV_1", "NAZIONE_1", "PRESSO",
    "TOPONIMO_2", "VIA_2", "CIV_2", "CAP_2", "ISTAT_2", "LOCALITA_2", "PROV_2", "NAZIONE_2",
    "TIPO_CONTRATTO", "SETT_MERCEOLOGICO",
    "TRATTAMENTO_IVA", "STAG_RIC", "DATA_INIZIO", "DATA_FINE",
    "SOLLEV_PERSONE", "AUTOCERT_SOLL_PERS", "AUTOCERT_CONTR_CONNESSIONE",
    "SERVIZIO_CURVE_CARICO", "MAND_CONN", "DISALIMENTABILE",
    "CATEGORIA_DISALIMENTABILITA",
    "AUTOCERT_ACQUISIZIONE_CERTIFICAZIONE_ASL",
    "TEL_CELL_PREAVVISO_PERSONALIZZATO_PESSE",
    "AUTOCERT_LIBERATORIA_MANCATO_AVVISO_PESSE",
    "NOTE", "CATEGORIA_CLIENTE", "DA_ESEGUIRE_NON_PRIMA_DEL",
]

# Mappa colonna logica -> candidati varianti per find_column
S01_INPUT_COLUMNS: dict[str, list[str]] = {
    "CodiceVenditore": ["CodiceVenditore", "Codice Venditore", "COD_VENDITORE", "CODICE VENDITORE"],
    "POD": ["POD"],
    "COGNOME": ["COGNOME"],
    "NOME": ["NOME"],
    "RAGSOC": ["RAGSOC", "RAG_SOC", "RAGIONE SOCIALE"],
    "CF": ["CF", "CODICE FISCALE"],
    "PIVA": ["PIVA", "P. IVA", "P.IVA", "PARTITA IVA"],
    "TELREFPRAT": ["TELREFPRAT", "TELEFONO", "TEL"],
    "USO": ["USO", "TIPO USO"],
}


def _resolve_s01_columns(df: pd.DataFrame) -> tuple[dict[str, str | None], list[str]]:
    """Risolve i nomi reali delle colonne nel DataFrame, accumulando warnings."""
    resolved: dict[str, str | None] = {}
    warnings: list[str] = []
    for logical, candidates in S01_INPUT_COLUMNS.items():
        col = find_column(df, candidates, mode="exact")
        if col is None:
            col = find_column(df, candidates, mode="substring")
        resolved[logical] = col
        if col is None:
            warnings.append(f"Colonna '{logical}' non trovata (cercati: {', '.join(candidates)})")
    return resolved, warnings


def _build_s01_row(
    df: pd.DataFrame,
    row_idx: int,
    resolved: dict[str, str | None],
    uso_warnings: set[str],
) -> list[str]:
    """Costruisce una riga S01 a 50 colonne nello stesso ordine di S01_HEADER."""

    def get(logical: str) -> str:
        col = resolved.get(logical)
        if col is None:
            return ""
        return _clean_val(df[col].iloc[row_idx])

    uso_raw = get("USO").upper()
    if uso_raw in S01_USO_MAP:
        tipo_contratto, sett_merceologico = S01_USO_MAP[uso_raw]
    else:
        tipo_contratto, sett_merceologico = "", ""
        if uso_raw:
            uso_warnings.add(f"USO non mappato: '{uso_raw}'")

    return [
        # Header fissi (5)
        *S01_HEADER_FISSI,
        # Input (8)
        get("CodiceVenditore"),
        get("POD"),
        get("COGNOME"),
        get("NOME"),
        get("RAGSOC"),
        get("CF"),
        get("PIVA"),
        get("TELREFPRAT"),
        # Presenza cliente
        "SI",
        # Indirizzo 1 (9)
        "", "", "", "", "", "", "", "", "",
        # Indirizzo 2 (8)
        "", "", "", "", "", "", "", "",
        # Codici (2)
        tipo_contratto,
        sett_merceologico,
        # Extra date (4)
        "", "", "", "",
        # Sollev/Autocert (3)
        "NO", "NO", "01",
        # Servizio/Mand/Disalim (3)
        "", "SI", "SI",
        # Trailing (7)
        "", "", "", "", "", "", "",
    ]


def genera_s01_massivo(file_path: Path, upload_dir: Path) -> dict:
    """Genera CSV + XLSX nel formato S01 Massivo dal file Excel pratiche.

    Args:
        file_path: Path al file Excel sorgente.
        upload_dir: Directory dove salvare CSV e XLSX di output.

    Returns:
        dict con keys: job_id, csv_path, xlsx_path, rows_created, columns,
        rows_preview, warnings.
    """
    logger.info("Genera S01 Massivo: lettura %s", file_path.name)

    df = pd.read_excel(
        file_path,
        dtype=str,
        keep_default_na=False,
        engine="openpyxl",
    )
    df.columns = df.columns.str.strip()

    if len(df) == 0:
        raise ValueError("Il file non contiene righe di dati")

    resolved, warnings = _resolve_s01_columns(df)

    uso_warnings: set[str] = set()
    rows: list[list[str]] = []
    for i in range(len(df)):
        row = _build_s01_row(df, i, resolved, uso_warnings)
        # Scarta righe completamente vuote nei campi input (no POD, no anagrafica)
        input_values = row[5:13]
        if not any(v for v in input_values):
            continue
        rows.append(row)

    if not rows:
        raise ValueError("Nessuna riga valida trovata nel file")

    all_warnings = sorted(set(warnings) | uso_warnings)

    job_id = str(uuid.uuid4())
    csv_path = upload_dir / f"s01_massivo_{job_id}.csv"
    xlsx_path = upload_dir / f"s01_massivo_{job_id}.xlsx"

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(S01_HEADER)
        writer.writerows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "S01"
    ws.append(S01_HEADER)
    for r in rows:
        ws.append(r)
    wb.save(xlsx_path)
    wb.close()

    logger.info("Genera S01 Massivo: %d righe -> %s, %s", len(rows), csv_path.name, xlsx_path.name)

    return {
        "job_id": job_id,
        "csv_path": str(csv_path),
        "xlsx_path": str(xlsx_path),
        "rows_created": len(rows),
        "columns": S01_HEADER,
        "rows_preview": rows,
        "warnings": all_warnings,
    }
